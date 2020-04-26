#handle and store data from xbee here
import random
from datetime import datetime
import msgpack
from openpyxl import load_workbook
import serial
import time
import sys
import pickle





class Info(object):
	def __init__(self):
		self.workbook = load_workbook(filename="data_label.xlsx")
		self.ws = self.workbook.active
		self.column = (self.ws['G'])[1:]
		self.labels = {}
	
		#identifies array label and adds value to it appropriately
		for i in self.column:
			if (i.value not in self.labels) and (i.value is not None):
				self.labels[i.value] = 1
			elif (i.value is not None):
				self.labels[i.value] += 1


		self.mph = random.randint(10,20);
		self.rpm = random.randint(10,20),
		self.miles = random.randint(10,20);
		self.socTime = str(datetime.now().strftime("%H:%M:%S")),
		self.socVal = random.randint(10,200);


	def to_json(self):
		
		msgpack_data = {}
		
		for label in self.labels:
			num = self.labels[label]
			if label == "b" or label == "k":			
				array_num = []			
				for i in range(num):
					array_num.append(random.randint(10,20))
							
				msgpack_data[label] = array_num
			else:
				array_bool = []			
				for i in range(num):
					array_bool.append(random.randint(0,1))		
				msgpack_data[label] = array_bool

		return msgpack.packb(msgpack_data, use_bin_type=False)


	def output(self, port):
		ser = serial.Serial(port, 115200, timeout=1)

		while True:
			d = self.to_json()
			pickled = pickle.dumps(d)
						
			print(type(d))
			ser.write(d)
			print(pickled)
			


			#print(ser.send_break)
			#print(sys.getsizeof(d))
			unpickled = pickle.loads(pickled)
			dict_pr = msgpack.unpackb(unpickled)
			
			print(type(dict_pr))
			print(dict_pr[b'b'])
			time.sleep(.5)
			print("sent")


if __name__ == "__main__":
	serial_port = 'ttyS10'
	k = Info()
	k.output(serial_port)
